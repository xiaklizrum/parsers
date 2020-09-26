import json
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from parsers.functions import *


sheet_name = 'Locale'
locale = {}

FileList = [
    '/tmp/parsers/jsontoxlsx/files/input/ru.json',
    '/tmp/parsers/jsontoxlsx/files/input/en.json',
    '/tmp/parsers/jsontoxlsx/files/input/kr.json',
    '/tmp/parsers/jsontoxlsx/files/input/zh.json',
]


def load_locale(path):
    locale_json = open(path, 'r')
    body = locale_json.read()
    return json.loads(body)


def search_by_key(key, locale_id):
    try:
        return str(locale[FileList[locale_id]]['common'][key])
    except KeyError:
        return ''


def json_to_xlsx():
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name)
    for index, (key, value) in enumerate(locale[FileList[0]]['common'].items()):
        ws.append([
            key, 
            str(value), 
            search_by_key(key, 1),
            search_by_key(key, 2), 
            search_by_key(key, 3)
        ])
    wb.save('/tmp/parsers/jsontoxlsx/files/output/output.xlsx')


def xlsx_to_json():
    _dict = {'common': {}}
    wb = load_workbook(filename='input.xlsx', read_only=True)
    ws = wb[sheet_name]
    for row in range(1, ws.max_row):
        if not ws['A' + row].value:
            break
        try:
            _dict['common'][ws['A' + index]] = str(ws['D' + row].value)
        except TypeError:
            pass
    file = open('/tmp/parsers/jsontoxlsx/files/output/output.json', 'w')
    file.write(escape(_dict))
    file.close()


if __name__ == '__main__':
    for locale_json in FileList:
        locale[locale_json] = load_locale(locale_json)
    try:
        import_type = sys.argv[1] 
    except IndexError:
        print('Enter one of two param: json2xlsx or xlsx2json')
    if import_type == 'json2xlsx':
        json_to_xlsx()
    elif import_type == 'xlsx2json':
        xlsx_to_json()
    else:
        print('Enter one of two param: json2xlsx or xlsx2json')
