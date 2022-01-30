import json
import sys
from openpyxl import Workbook, load_workbook
from os import listdir
from os.path import isfile, join
from parsers.functions import escape


INPUT_DIR = "/tmp/parsers/json2xlsx/files/input/"
OUTPUT_DIR = "/tmp/parsers/json2xlsx/files/output/"

files = {
    "input": {
        "json": [],
        "xlsx": [],
    }
}

files_content = {}


def load_file(file):
    with open(INPUT_DIR + file, "r") as file:
        content = file.read()
    return json.loads(content)


def search_by_key(key, filename):
    try:
        return str(files_content[filename]["main"][key])
    except KeyError:
        return ""


def json_to_xlsx():
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    title = ["key"]
    title.extend(list(files_content.keys()))
    ws.append(title)
    for key in files_content[title[1]]["main"].keys():
        row = [key]
        for file in title[1:]:
            row.append(search_by_key(key, file))
        ws.append(row)
    wb.save(OUTPUT_DIR + "output.xlsx")


def xlsx_to_json():
    _dict = {"main": {}}
    wb = load_workbook(filename="{}input.xlsx".format(INPUT_DIR))
    ws = wb.worksheets[0]

    files = [x.value for x in list(ws.rows)[0]][1:]

    for column, file in enumerate(files):
        for index in range(2, ws.max_row):
            key = ws.cell(column=1, row=index).value
            if not key:
                break
            try:
                value = ws.cell(column=column + 2, row=index).value
                _dict["main"][key] = value
            except TypeError:
                pass
        file = open("{}{}.json".format(OUTPUT_DIR, file), "w")
        file.write(escape(_dict))
        file.close()


if __name__ == "__main__":
    inputfiles = [f for f in listdir(INPUT_DIR) if isfile(join(INPUT_DIR, f))]

    for file in inputfiles:
        if ".json" in file:
            files["input"]["json"].append(file)
        elif ".xlsx" in file:
            files["input"]["xlsx"].append(file)
    try:
        import_type = sys.argv[1]
    except IndexError:
        print("Enter one of two param: json2xlsx or xlsx2json")

    if import_type == "json2xlsx":
        input_file_format = "json" if import_type == "json2xlsx" else "xlsx"
        for file in files["input"][input_file_format]:
            files_content["".join(file.split(".")[:1])] = load_file(file)
        json_to_xlsx()
    elif import_type == "xlsx2json":
        xlsx_to_json()
    else:
        print("Enter one of two param: json2xlsx or xlsx2json")
